import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def add_table_sheet(wb: Workbook, title: str, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return

    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_overview(wb: Workbook, by_group_rows: list[list[str]], extra_rows: list[list[str]]) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "LLM辅助Linux内核缺陷修复 - 结项结果总览"
    ws["A1"].font = Font(bold=True, size=16)

    overview = [
        ["指标", "结果"],
        ["真实 kBenchSyz 样本数", "8"],
        ["实验分组", "Baseline / With Trace / Improved"],
        ["模型", "DeepSeek V4 Pro"],
        ["模型输出数", "24"],
        ["JSON 解析失败", "0"],
        ["Plausible", "2/24 (8.3%)"],
        ["Helpful", "7/24 (29.2%)"],
        ["Incorrect", "15/24 (62.5%)"],
        ["Plausible + Helpful", "9/24 (37.5%)"],
        ["说明", "未进行完整 Linux kernel build / reproducer 动态验证"],
    ]
    for row in overview:
        ws.append(row)

    ws["A14"] = "按实验组统计"
    ws["A14"].font = Font(bold=True, size=13)
    start = 15
    for r, row in enumerate(by_group_rows, start=start):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row_idx in [2, start]:
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in range(1, 11):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    if len(by_group_rows) >= 2:
        chart = BarChart()
        chart.title = "Plausible + Helpful Rate by Group"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Group"
        data = Reference(ws, min_col=9, min_row=start, max_row=start + len(by_group_rows) - 1)
        cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(by_group_rows) - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, "A22")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(ROOT / "results" / "evaluation.xlsx"))
    args = parser.parse_args()

    sources = {
        "Samples": ROOT / "data" / "selected_bugs.csv",
        "Manual Evaluation": ROOT / "results" / "evaluation_real.csv",
        "Group Summary": ROOT / "results" / "evaluation_summary_by_group.csv",
        "Patch Checks": ROOT / "results" / "check_results_summary_deepseek.csv",
        "Kernel Verify": ROOT / "results" / "kernel_verify_summary.csv",
        "Compile Compare": ROOT / "results" / "local_compile_comparison" / "summary.csv",
    }
    loaded = {name: read_csv(path) for name, path in sources.items()}

    wb = Workbook()
    build_overview(wb, loaded["Group Summary"], [])
    for name, rows in loaded.items():
        add_table_sheet(wb, name, rows)

    local_compile_path = ROOT / "results" / "local_compile" / "summary.json"
    if local_compile_path.exists():
        summary = json.loads(local_compile_path.read_text(encoding="utf-8"))
        rows = [["field", "value"]] + [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value] for key, value in summary.items()]
        add_table_sheet(wb, "Local Compile", rows)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
